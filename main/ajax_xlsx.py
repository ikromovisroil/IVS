from io import BytesIO
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import *


@require_GET
@login_required
def export_technics_xlsx(request):
    org_id = (request.GET.get("organization") or "").strip()
    category_id = (request.GET.get("category") or "").strip()
    status = (request.GET.get("status") or "").strip()
    name = (request.GET.get("name") or "").strip()

    qs = (
        Technics.objects
        .all()
        .select_related("organization", "category")
        .order_by("-id")
    )

    if org_id.isdigit():
        qs = qs.filter(organization_id=int(org_id))

    if category_id.isdigit():
        qs = qs.filter(category_id=int(category_id))

    if status:
        qs = qs.filter(status=status)

    if name:
        qs = qs.filter(name__icontains=name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Technics"

    headers = ["Category", "Name", "Parametr", "I/N", "S/N", "Mac", "Status"]
    ws.append(headers)

    for t in qs:
        ws.append([
            (t.category.name if getattr(t, "category", None) else ""),
            (getattr(t, "name", "") or ""),
            (getattr(t, "parametr", "") or ""),
            (getattr(t, "inventory", "") or ""),
            (getattr(t, "serial", "") or ""),
            (getattr(t, "mac", "") or ""),
            (getattr(t, "status", "") or ""),
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "technics.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@require_GET
@login_required
def export_material_xlsx(request):
    status = (request.GET.get("status") or "").strip()
    employee_id = (request.GET.get("employee") or "").strip()
    name = (request.GET.get("name") or "").strip()

    qs = (
        Material.objects
        .all()
        .select_related("employee")   # agar Material.employee FK bo'lsa
        .order_by("-id")
    )

    if status:
        qs = qs.filter(status=status)

    # Material biriktirilgan xodim bo'yicha filter
    if employee_id.isdigit():
        qs = qs.filter(employee_id=int(employee_id))

    if name:
        qs = qs.filter(name__icontains=name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Material"

    headers = ["Xodim", "Material Nomi", "Soni", "Narxi", "Qiymati", "1C code", "Status"]
    ws.append(headers)

    for m in qs:
        emp = getattr(m, "employee", None)
        emp_name = ""
        if emp:
            # sizda full_name yo'q bo'lishi mumkin, shuning uchun safe:
            emp_name = " ".join(filter(None, [
                getattr(emp, "last_name", ""),
                getattr(emp, "first_name", ""),
                getattr(emp, "father_name", ""),
            ])).strip()

        ws.append([
            emp_name,
            getattr(m, "name", "") or "",
            getattr(m, "number", "") or "",
            getattr(m, "price", "") or "",
            (m.unit.name if getattr(m, "unit", None) else ""),
            getattr(m, "code", "") or "",
            getattr(m, "status", "") or "",
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "material.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp