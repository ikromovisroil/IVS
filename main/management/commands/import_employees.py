import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from django.db import transaction
from django.utils.text import slugify

from main.models import (
    Employee,
    Organization,
    Department,
    Region,
    Rank,
)
from main.services.fio_split import split_fio  # sendagi funksiya


class Command(BaseCommand):
    help = (
        "Shtat jadvalidan xodimlarni import qilish: "
        "avval Employee bor-yo'qligini FIO bo'yicha tekshiradi, "
        "bo'lsa faqat o'shani yangilaydi, "
        "bo'lmasa yangi User + Employee (yoki signaldan kelgan Employee) ni to'ldiradi."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Excel fayl yo'li, masalan: import/Shtat.xlsx",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="01.05.2025",  # kerak bo'lsa o'zgartirasan
            help="Varaq nomi (default: 01.05.2025)",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = options["file_path"]
        sheet_name = options["sheet"]

        # 1) Tashkilot (IVS)
        try:
            org = Organization.objects.get(org_type="IVS")
        except Organization.DoesNotExist:
            self.stderr.write(self.style.ERROR("❌ org_type='IVS' bo'lgan Organization topilmadi"))
            return

        # 2) Region – hozircha 'Toshkent'
        region, _ = Region.objects.get_or_create(name="Toshkent")

        # 3) Excelni ochish
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            self.stderr.write(
                self.style.ERROR(
                    f"❌ '{sheet_name}' varaq topilmadi. Mavjud varaqalar: {wb.sheetnames}"
                )
            )
            return

        ws = wb[sheet_name]

        created = 0
        updated = 0
        skipped = 0

        # 1-qator — sarlavha
        for row in ws.iter_rows(min_row=2, values_only=True):
            # ⚠️ Shtat.xlsx: LAVOZIM | F.I.O | DEP | ...
            lavozim = (row[0] or "").strip() if row[0] else ""
            fio_raw = (row[1] or "").strip() if row[1] else ""
            dep_name = (row[2] or "").strip() if len(row) > 2 and row[2] else ""

            if not fio_raw:
                skipped += 1
                continue

            # FIO ni bo'lamiz
            last_name, first_name, father_name = split_fio(fio_raw)

            if not last_name or not first_name:
                skipped += 1
                continue

            # Department
            department = None
            if dep_name:
                department, _ = Department.objects.get_or_create(
                    organization=org,
                    name=dep_name,
                )

            # Rank (lavozim)
            rank = None
            if lavozim:
                rank, _ = Rank.objects.get_or_create(name=lavozim)

            # 4) AVVAL FIO + IVS bo'yicha Employee izlaymiz
            emp = (
                Employee.objects.filter(
                    organization=org,
                    last_name__iexact=last_name,
                    first_name__iexact=first_name,
                    father_name__iexact=father_name,
                )
                .select_related("user")
                .first()
            )

            # -------------------------------
            #  A) EMPLOYEE MAVJUD
            # -------------------------------
            if emp:
                self._update_employee(
                    emp=emp,
                    org=org,
                    region=region,
                    department=department,
                    rank=rank,
                )
                updated += 1
                continue

            # -------------------------------
            #  B) EMPLOYEE YO'Q → YANGI USER
            # -------------------------------
            base_username = slugify(f"{last_name}.{first_name}") or "user"
            username = base_username
            i = 1
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{i}"
                i += 1

            user = User.objects.create_user(
                username=username,
                first_name=first_name,
                last_name=last_name,
                password="Password100",
            )

            # ❗ MUHIM JOY: SIGNALLAR NATIJASIDA ALREADY Employee YARATILGAN-BO'LMAGANINI TEKSHIRAMIZ
            emp_for_user = None
            try:
                emp_for_user = user.employee  # agar OneToOne signal ishlagan bo'lsa, bu bor bo'ladi
            except Employee.DoesNotExist:
                emp_for_user = None
            except AttributeError:
                emp_for_user = None

            if emp_for_user:
                # Signal orqali yaratilgan Employee ni to'ldiramiz
                emp_obj = emp_for_user
                emp_obj.organization = org
                emp_obj.region = region
                emp_obj.last_name = last_name
                emp_obj.first_name = first_name
                emp_obj.father_name = father_name
                emp_obj.department = department
                emp_obj.rank = rank
                emp_obj.status = "worker"
                emp_obj.save()
                updated += 1
            else:
                # Signal Employee yaratmagan bo'lsa, endi o'zimiz yaratamiz
                Employee.objects.create(
                    user=user,
                    organization=org,
                    region=region,
                    department=department,
                    rank=rank,
                    last_name=last_name,
                    first_name=first_name,
                    father_name=father_name,
                    status="worker",
                )
                created += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"✅ Import yakunlandi: Yaratildi: {created} | Yangilandi: {updated} | O'tkazib yuborildi: {skipped}"
            )
        )

    def _update_employee(self, emp, org, region, department, rank):
        """
        Mavjud Employee obyektini yangilash.
        """
        emp.organization = org
        emp.region = emp.region or region

        if department:
            emp.department = department  # save() ichida organization ham to'ladi
        if rank:
            emp.rank = rank

        if not emp.status or emp.status == "client":
            emp.status = "worker"

        emp.save()
